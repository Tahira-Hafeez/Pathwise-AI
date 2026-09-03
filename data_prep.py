

import json
import openpyxl

SOURCE_FILE = "data/PathWise_AI_Knowledge_Base_milestone1.xlsx"
OUTPUT_FILE = "data/skills.json"


SKILL_SHEETS = ["AI Engineer", "Data Analyst"]


def load_skills():
    wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
    all_skills = []

    for sheet_name in SKILL_SHEETS:
        matched_sheet = next(
            (s for s in wb.sheetnames if s.strip() == sheet_name.strip()), None
        )
        if not matched_sheet:
            print(f"Warning: sheet '{sheet_name}' not found, skipping")
            continue

        ws = wb[matched_sheet]
        rows = list(ws.iter_rows(values_only=True))

        role_title = str(rows[0][0]).strip()
        header_row = rows[1]
        seen_ids = set()
        for row in rows[2:]:
            if not row[0]:
                continue
            if str(row[0]).strip().lower() in ("skill id",):
                continue
            skill_id, skill, prereq, why, resource, est_time = row[:6]
            all_skills.append({
                "role": role_title,
                "skill_id": skill_id,
                "skill": skill.strip() if skill else "",
                "prerequisites": prereq if prereq else "None",
                "why_it_matters": why if why else "",
                "resource": resource if resource else "",
                "est_time": est_time if est_time else "",
            })

    return all_skills


if __name__ == "__main__":
    skills = load_skills()
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(skills, f, indent=2, ensure_ascii=False)
    print(f"Saved {len(skills)} skill records to {OUTPUT_FILE}")
